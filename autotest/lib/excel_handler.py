# coding:utf-8
from openpyxl import load_workbook
import openpyxl.styles as style


class ExcelHandler:
    def __init__(self, filename):
        self.filename = filename
        self.workbook = self.open_file(filename)

    def open_file(self, filename=""):
        try:
            wk = load_workbook(filename=self.filename, read_only=False, data_only=True)
        except Exception as e:
            print("open excel(%s) fail:%s" % (self.filename, e))
            return None
        return wk

    def get_sheet(self, sheet_name):
        if self.workbook:
            if sheet_name == "":
                return self.workbook.active
            else:
                return self.workbook[sheet_name]
        return None

    def close(self):
        if self.workbook:
            self.workbook.close()

    def read_sheet_to_array(self, sheet_name=""):
        try:
            sheet = self.get_sheet(sheet_name)
            vlaues = sheet.values
            return vlaues
        except Exception as e:
            print("读取文件(%s)异常:%s" % (self.filename, e))
            return None

    def read_cell(self, row_index, column_index, sheet_name=""):
        try:
            sheet = self.get_sheet(sheet_name)
            return sheet.cell(row=row_index, column=column_index).value
        except Exception as e:
            print("读取文件(%s)cell数据(%s,%s)异常:%s" % (self.filename, row_index, column_index, e))
            return None

    def get_datavalidation_info(self, row_index, column_index, sheet_name=""):
        """
        获取指定单元格的数据有效性信息
        :param row_index:
        :param column_index:
        :param sheet_name:
        :return:
        """
        try:
            sheet = self.get_sheet(sheet_name)
            # 获取所有的验证数据信息
            dvs = sheet.data_validations.dataValidation
            match_dv = None
            is_matched = False
            # 匹配指定单元格上应用的验证数据信息
            for dv in dvs:
                ranges = dv.cells.ranges
                for rg in ranges:
                    if (rg.min_row <= row_index <= rg.max_row) and (rg.min_col <= column_index <= rg.max_col):
                        match_dv = dv
                        is_matched = True
                        break
                if is_matched:
                    break
            return match_dv.formula1

        except Exception as e:
            print("读取文件(%s)指定cell(%s,%s)有效数据失败:%s" % (self.filename, row_index, column_index, e))
            return None

    def write_sheet_by_df(self, dataset, rgb=None, sheet_name=""):
        """
        写入excel，并可以通过rgb参数设置背景色
        :param dataset:
        :param rgb:
        :param sheet_name:
        :return:
        """
        ws = self.get_sheet(sheet_name)
        for data in dataset:
            row = ws.max_row
            ws.append(data)

            if rgb:
                for i in range(1, len(data) + 1):
                    ws.cell(row=row + 1, column=i).fill = style.PatternFill(fill_type='solid', fgColor=rgb)

    def save(self):
        ws = self.get_sheet("")
        if not self.workbook.read_only:
            self.workbook.save(self.filename)
        else:
            print("文件(%s)只读，不能保存!" % self.filename)


if __name__ == '__main__':
    handler = ExcelHandler(u"D:\\temp\\doc\\测试\\测试脚本更新上传\\准生产存管三方记账.xlsx")
    handler.save()
    print(handler.get_datavalidation_info(1, 2))
